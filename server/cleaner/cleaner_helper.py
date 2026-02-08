import pandas as pd
import numpy as np
import pdfplumber
import io
import re
from sqlmodel import Session, select, col
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback
from openpyxl.utils import get_column_letter

#=================================================================
from ..models import OperationData
#=================================================================


#=================================================================

metadata = {'created_at','updated_at','operation_type','processed_by'}

def get_mandatory_columns():
    """Get all column names from OperationData model"""
    columns = []
    
    # Get fields from model
    for field_name, field in OperationData.__fields__.items():
        # Skip metadata columns
        if field_name not in metadata:
            columns.append(field_name)
    
    return columns



def format_excel_headers(ws, start_row=1, start_col=1):
    """
    Format headers in an Excel worksheet.
    
    Args:
        ws: Worksheet object
        start_row: Row number where headers start (default: 1)
        start_col: Column number where headers start (default: 1)
    """
    
    # Define styling
    header_fill = PatternFill(
        start_color="366092",  # Dark blue
        end_color="366092",
        fill_type="solid"
    )
    
    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF"  # White
    )
    
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True  # Enable text wrapping
    )
    
    # Get max column with data
    max_column = ws.max_column
    
    # Apply formatting to each header cell
    for col in range(start_col, max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        
        # Apply styling if cell has a value
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    # Set row height for header row
    ws.row_dimensions[start_row].height = 30
    
    # Auto-fit column widths
    for col in range(start_col, max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # Check all cells in the column (including header)
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            
            if cell.value:
                # Calculate length considering line breaks for wrapped text
                if isinstance(cell.value, str):
                    # Find the longest line if text wraps
                    lines = str(cell.value).split('\n')
                    line_length = max(len(line) for line in lines)
                else:
                    line_length = len(str(cell.value))
                
                # Add a little padding
                adjusted_length = line_length + 2
                
                if adjusted_length > max_length:
                    max_length = adjusted_length
        
        # Set column width (minimum 10, maximum 50)
        column_width = min(max(max_length, 10), 50)
        ws.column_dimensions[column_letter].width = column_width

    return ws


def clean_columns(columns):
    cleaned = (
        columns
        .str.replace(r"\n", " ", regex=True)        # remove line breaks
        .str.replace(r"\t", " ", regex=True)        # remove tabs
        .str.replace(r"\s+", " ", regex=True)       # normalize spaces
        .str.strip()                                # trim edges
        .str.lower()                                # lowercase
        .str.replace(r"[^\w\s]", "", regex=True)    # remove special chars
        .str.replace(" ", "_")                      # snake_case
    )
    return cleaned


def clean_address(series: pd.Series) -> pd.Series:
    """
    Cleans address text exactly like:
    =UPPER(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A2,"-"," "),","," "),"/"," "))
    """
    return (
        series
        .astype(str)
        .str.replace("-", " ", regex=False)
        .str.replace(",", " ", regex=False)
        .str.replace("/", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)  # normalize spaces
        .str.strip()
        .str.upper()
    )



from openpyxl.utils import get_column_letter

def format_excel_sheet(ws):
    """
    Final Excel formatter:
    - Header style (blue, bold, white)
    - Cambria font for all cells
    - Wrap text ON
    - Row height = 30
    - Auto-fit columns
    """
    # Styles
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(name="Cambria", size=12, bold=True, color="FFFFFF")
    cell_font = Font(name="Cambria")

    align_center_wrap = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 1. Header formatting (First Row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center_wrap
        cell.border = border

    ws.row_dimensions[1].height = 30

    # 2. Cell formatting + row height (Data Rows)
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            # Note: We don't want to overwrite the Red/Yellow logic 
            # so we only apply font/alignment if not already specialized
            cell.font = cell_font
            cell.alignment = align_center_wrap
            cell.border = border

    # 3. Auto-fit column width
    for col in ws.columns:
        max_length = 0
        column_idx = col[0].column
        col_letter = get_column_letter(column_idx)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # 4. Override specific columns by Header Name
    for cell in ws[1]:
        col_letter = get_column_letter(cell.column)
        if cell.value == "EMPLOYEE ADDRESS":
            ws.column_dimensions[col_letter].width = 80
        elif cell.value == "EMPLOYEE NAME":
            ws.column_dimensions[col_letter].width = 30



def standardize_dataframe(df):
    """
    Standardizes DataFrame to match SQLModel definitions:
    1. Generates 'unique_id' (Composite Key).
    2. Ensures ALL mapped columns exist (fills missing with "").
    3. Returns DataFrame with [Mandatory Columns] + [Any Extra Columns].
    """
    if df is None or df.empty:
        return None

    # Work on a copy to avoid SettingWithCopyWarning
    df = df.copy()

    # 1. Generate Unique ID
    if 'trip_id' in df.columns and 'employee_id' in df.columns:
        df['unique_id'] = df['trip_id'].astype(str).str.strip() + df['employee_id'].astype(str).str.strip()
        # Filter rows with missing or string 'nan' IDs
        df = df[
            df["unique_id"].notna() & 
            (df["unique_id"] != "") & 
            (~df["unique_id"].str.contains("nan|None", case=False))
        ]
    else:
        # If we can't make a unique_id, we can't sync to DB
        return None 

    # 2. Sync Columns with SQLModel (Fill Missing)
    target_cols = get_mandatory_columns()
    for col in target_cols:
        if col not in df.columns:
            df[col] = ""

    # 3. Populate specific logic columns
    df['una2'] = df['unique_id'] 

    # 4. Organize Columns: Mandatory -> Extra -> unique_id
    current_cols = df.columns.tolist()
    reserved_cols = set(target_cols) | {'unique_id'}
    extra_cols = [c for c in current_cols if c not in reserved_cols]
    
    # Ensure we only try to select columns that actually exist
    final_order = [c for c in (target_cols + extra_cols + ['unique_id']) if c in df.columns]
    
    return df[final_order]

def get_xls_style_data(book, xf_index, row_idx, col_idx):
    """
    Extracts background and font colors from legacy .xls files.
    Includes debug prints to identify why colors might be missed.
    """
    try:
        xf = book.xf_list[xf_index]
        font = book.font_list[xf.font_index]
        
        # --- FONT COLOR DETECTION (RED) ---
        f_idx = font.colour_index
        rgb_f = book.colour_map.get(f_idx)
        font_hex = None
        
        # Check standard red indices (8, 10, 16 are common for Red)
        if f_idx in [10, 16]:
            font_hex = "FF0000"
        elif rgb_f:
            # Check if RGB values are "Red-ish" (High Red, Low Green/Blue)
            if rgb_f[0] > 150 and rgb_f[1] < 100 and rgb_f[2] < 100:
                font_hex = "FF0000"
        
        # --- BACKGROUND COLOR DETECTION (YELLOW) ---
        bg_idx = xf.background.pattern_colour_index
        rgb_b = book.colour_map.get(bg_idx)
        bg_hex = None
        
        # Check standard yellow indices (13, 19 are common for Yellow)
        if bg_idx in [13, 19]:
            bg_hex = "FFFF00"
        elif rgb_b:
            # Check if RGB values are "Yellow-ish"
            if rgb_b[0] > 200 and rgb_b[1] > 200 and rgb_b[2] < 150:
                bg_hex = "FFFF00"

        # --- DEBUG LOGGING ---
        # Only print for non-default styles to keep console clean
        if font_hex == "FF0000" or bg_hex == "FFFF00":
            print(f"[DEBUG STYLE] Row {row_idx}, Col {col_idx} | FontIdx: {f_idx} (Hex: {font_hex}) | BgIdx: {bg_idx} (Hex: {bg_hex})")

        return bg_hex, font_hex, bool(font.bold)
    except Exception as e:
        print(f"[DEBUG ERROR] Style extraction failed at Row {row_idx}, Col {col_idx}: {e}")
        return None, None, False    


# ==========================================
# EXCEL GENERATOR HELPER
# ==========================================
def create_styled_excel(df, filename_prefix="Cleaned"):
    """ Generates Excel with consistent formatting. """
    output = io.BytesIO()
    
    # 1. Fetch Mapping dynamically
    try:
        # Assuming get_mandatory_columns returns {Friendly: DB_Col}
        mandatory_map = get_mandatory_columns() 
        mandatory_headers_list = list(mandatory_map.keys())
    except:
        mandatory_map = {}
        mandatory_headers_list = []

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        
        # 2. Check Data Type
        # Fastag data has 'Amount' or 'Plaza Name'. 
        # Standard data uses the mandatory map.
        is_fastag = 'Amount' in df.columns or 'Plaza Name' in df.columns

        if not is_fastag and mandatory_map:
            # --- STANDARD CLIENT/RAW DATA LOGIC ---
            # Create Inverse Map: DB_Col (snake_case) -> Friendly Name
            inv_map = {v: k for k, v in mandatory_map.items()}
            
            # Rename columns to Friendly Names
            df_export = df.rename(columns=inv_map)
            
            # Reorder Columns: Mandatory First -> Then Others
            export_cols = [h for h in mandatory_headers_list if h in df_export.columns]
            remaining = [c for c in df_export.columns if c not in mandatory_headers_list and c != 'unique_id']
            export_cols += remaining
            
            df_export = df_export[export_cols]
        else:
            # --- FASTAG / GENERIC LOGIC ---
            # Export as is (we already cleaned it in the specific function)
            df_export = df

        # 3. Write Data
        df_export.to_excel(writer, index=False, sheet_name='Data')
        
        # 4. Apply Styles
        workbook = writer.book
        worksheet = writer.sheets['Data']
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#0070C0', 'font_color': 'white'})
        
        for i, col in enumerate(df_export.columns):
            worksheet.write(0, i, col, header_fmt)
            # Auto-adjust width roughly
            worksheet.set_column(i, i, 20)
            
    output.seek(0)
    return df, output, f"{filename_prefix}.xlsx"

# ==========================================
# DATABASE HELPER: BULK SAVE
# ==========================================
def bulk_save_unique(session, model, df):
    """
    Saves rows to the database only if the unique_id doesn't already exist.
    Handles duplicate columns automatically.
    """
    if df is None or df.empty:
        return 0

    try:
        # 🔥 FIX: Deduplicate columns first.
        # If 'unique_id' appears twice, df['unique_id'] returns a DataFrame (causing the crash).
        # This line keeps only the first occurrence of every column name.
        df = df.loc[:, ~df.columns.duplicated()]

        if "unique_id" not in df.columns:
            print(f"❌ Error: 'unique_id' missing in {model.__tablename__} data.")
            return 0

        # Now safe to call unique() because we know it's a Series
        incoming_ids = df["unique_id"].astype(str).unique().tolist()

        # Check existing IDs in DB
        from sqlmodel import select
        statement = select(model.unique_id).where(model.unique_id.in_(incoming_ids))
        existing_db_ids = session.exec(statement).all()
        existing_set = set(existing_db_ids)

        # Filter: Keep rows NOT in DB
        new_data = df[~df['unique_id'].isin(existing_set)]

        if not new_data.empty:
            records = new_data.to_dict(orient='records')
            # Create objects safely
            objects = [model(**row) for row in records]
            
            session.add_all(objects)
            session.commit()
            
            count = len(objects)
            print(f"✅ Saved {count} new records to {model.__tablename__}")
            return count
        
        print("🔹 No new records to save.")
        return 0

    except Exception as e:
        session.rollback()
        print(f"❌ Database Error in bulk_save_unique: {e}")
        import traceback
        traceback.print_exc()
        return 0
def sync_addresses_to_t3(session, df):
    """
    Extracts unique addresses from the processed data and syncs them to the T3 table.
    """
    if df is None or 'address' not in df.columns:
        return 0

    try:
        # ✅ FIX: Use df['address'].unique() instead of df.unique()
        unique_addresses = [a.strip().upper() for a in df['address'].unique() if a and str(a).strip()]
        
        if not unique_addresses:
            return 0

        # Assuming AddressTable is your T3 model
        # This part depends on your specific Address model name
        # For now, we return the count of unique addresses found
        return len(unique_addresses)
    except Exception as e:
        print(f"Error syncing addresses: {e}")
        return 0
import pandas as pd
import numpy as np
import pdfplumber
import io
import re
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import traceback
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
import pandas as pd
from datetime import datetime
from io import BytesIO

from .cleaner_helper import (
    get_mandatory_columns, 
    get_xls_style_data, 
    standardize_dataframe,
    format_excel_sheet,
    clean_columns,
    clean_address,
    
)


# ==========================================
# 1. APP OPERATION DATA CLEANER
# ==========================================


def process_operation_app_data(file_list_bytes):
    # 1. Configuration
    COLUMN_TO_RENAME = {
        'DATE': 'shift_date', 'TRIP ID': 'trip_id', 'FLT NO.': 'flight_number', 
        'SAP ID': 'employee_id', 'EMP NAME': 'employee_name', 'EMPLOYEE ADDRESS': 'address', 
        'PICKUP LOCATION': 'landmark', 'DROP LOCATION': 'office', 'CAB NO': 'cab_last_digit',
        'PICKUP TIME': 'pickup_time', 'REMARKS': 'mis_remark'
    }
    SKIP_HEADERS = ['CONTACT NO', 'GUARD ROUTE', 'AIRPORT DROP TIME']

    wb = Workbook()
    ws = wb.active
    ws.title = "Operation_Data"
    
    # ... (Styles setup same as before) ...
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    try:
        MANDATORY_HEADERS = get_mandatory_columns()
    except:
        MANDATORY_HEADERS = list(COLUMN_TO_RENAME.values())

    for col_idx, header in enumerate(MANDATORY_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    target_row = 2
    data_rows = [] 
    extra_headers_map = {} 
    next_extra_col_idx = len(MANDATORY_HEADERS) + 1

    # 3. Processing Loop
    for filename, content in file_list_bytes:
        if not filename.lower().endswith('.xls'): continue
        print(f"\n--- Processing File: {filename} ---")

        try:
            rb = xlrd.open_workbook(file_contents=content, formatting_info=True)
            rs = rb.sheet_by_index(0)
            source_headers = [str(rs.cell_value(0, c)).strip().upper() for c in range(rs.ncols)]
            
            # Map columns
            # --- START ADDED LOGIC: IDENTIFY SPECIFIC COLUMN INDICES ---
            idx_trip = next((i for i, h in enumerate(source_headers) if 'TRIP ID' in h), None)
            idx_sap = next((i for i, h in enumerate(source_headers) if 'SAP ID' in h), None)
            idx_addr = next((i for i, h in enumerate(source_headers) if 'EMPLOYEE ADDRESS' in h), None)

            col_to_target_map = {}
            # --- END ADDED LOGIC ---

            for idx, raw_header in enumerate(source_headers):
                if any(skip in raw_header for skip in SKIP_HEADERS): continue
                match = next((val for key, val in COLUMN_TO_RENAME.items() if key in raw_header), None)
                if match:
                    col_to_target_map[idx] = {'type': 'mandatory', 'name': match}
                else:
                    if raw_header not in extra_headers_map:
                        extra_headers_map[raw_header] = next_extra_col_idx
                        ws.cell(row=1, column=next_extra_col_idx, value=raw_header)
                        next_extra_col_idx += 1
                    col_to_target_map[idx] = {'type': 'extra', 'name': raw_header}

            # 4. Process Data Rows
            for r_idx in range(1, rs.nrows):
                row_vals = [str(rs.cell_value(r_idx, c)).strip() for c in range(rs.ncols)]
                if sum(1 for v in row_vals if v != "") <= 3: 
                    continue

                row_data_map = {} 
                db_row_dict = {}
                
                # Logic Flags for Style Detection across the entire row
                row_has_yellow_bg = False
                row_has_red_font = False

                # --- START ADDED LOGIC: 3-COLUMN COUNTERS ---
                red_count = 0
                yellow_count = 0
                check_indices = [idx for idx in [idx_trip, idx_sap, idx_addr] if idx is not None]
                # --- END ADDED LOGIC ---

                # Pass 1: Extract data and scan row for color indicators
                for c_idx in range(rs.ncols):
                    bg, fg, is_bold = get_xls_style_data(rb, rs.cell_xf_index(r_idx, c_idx), r_idx, c_idx)
                    
                    # --- START ADDED LOGIC: UPDATE COUNTERS BASED ON 3 SPECIFIC COLUMNS ---
                    if c_idx in check_indices:
                        if fg == "FF0000": red_count += 1
                        if bg == "FFFF00": yellow_count += 1
                    # --- END ADDED LOGIC ---
                    
                    if fg == "FF0000": row_has_red_font = True
                    if bg == "FFFF00": row_has_yellow_bg = True
                    
                    if c_idx in col_to_target_map:
                        target_header = col_to_target_map[c_idx]['name']
                        val = rs.cell_value(r_idx, c_idx)
                        row_data_map[target_header] = {'val': val, 'bg': bg, 'fg': fg, 'bold': is_bold}
                        db_row_dict[target_header] = val

                # --- START ADDED LOGIC: RE-EVALUATE FLAGS BASED ON 3-COLUMN RULE ---
                row_has_red_font = (red_count == 3)
                row_has_yellow_bg = (yellow_count == 3)
                # --- END ADDED LOGIC ---

                # Pass 2: Apply Business Logic Overrides (Priority: Red > Yellow)
                if row_has_red_font:
                    db_row_dict['mis_remark'] = "Cancel"
                    print(f"[LOGIC] Row {r_idx}: Red found -> Marked Cancel")
                elif row_has_yellow_bg:
                    db_row_dict['mis_remark'] = "Alt Veh"
                    print(f"[LOGIC] Row {r_idx}: Yellow found -> Marked Alt Veh")

                # Pass 3: Write to Excel Output
                ws.row_dimensions[target_row].height = 25
                for c_out, m_header in enumerate(MANDATORY_HEADERS, 1):
                    cell = ws.cell(row=target_row, column=c_out)
                    
                    # Get value from dict (contains overrides like 'Cancel')
                    cell.value = db_row_dict.get(m_header, "")
                    
                    # Styling for the mis_remark column based on triggers
                    if m_header == 'mis_remark':
                        if row_has_red_font:
                            cell.font = Font(color="FF0000", bold=True)
                        elif row_has_yellow_bg:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
                            cell.font = Font(bold=True)
                    elif m_header in row_data_map:
                        # Carry over original formatting for other columns
                        d = row_data_map[m_header]
                        if d['bg']:
                            cell.fill = PatternFill(start_color=d['bg'], end_color=d['bg'], fill_type='solid')
                        cell.font = Font(bold=d['bold'], color=d['fg'] if d['fg'] else None)
                    
                    cell.alignment = align_center
                    cell.border = border

                # Extra check for dynamic extra columns
                for extra_name, extra_col_idx in extra_headers_map.items():
                    cell = ws.cell(row=target_row, column=extra_col_idx)
                    if extra_name in row_data_map:
                        cell.value = row_data_map[extra_name]['val']
                        cell.alignment = align_center
                        cell.border = border

                # Append row to DB list
                if db_row_dict.get('employee_id') or db_row_dict.get('employee_name'):
                    data_rows.append(db_row_dict)
                    target_row += 1
            
            rb.release_resources()
        except Exception as e:
            print(f"[BREAKING ERROR] File {filename}: {e}")
            traceback.print_exc()

    # --- DATAFRAME POST-PROCESSING ---
    df_db = pd.DataFrame(data_rows)
    if not df_db.empty:
        # 1. Helper: Convert Serial Date to DD-MM-YYYY
        def convert_date(d):
            try:
                # Handle Excel Serial (e.g., 46023)
                f_val = float(d)
                # Excel's base date is 1899-12-30
                dt = datetime(1899, 12, 30) + timedelta(days=f_val)
                return dt.strftime('%d-%m-%Y')
            except:
                return str(d)

        # 2. Helper: Convert Serial Time to HH:MM
        def convert_time(t):
            try:
                f_val = float(t) % 1 # MOD 1 logic
                seconds = int(round(f_val * 86400))
                return (datetime.min + timedelta(seconds=seconds)).strftime('%H:%M')
            except:
                return str(t)

        print("[DEBUG] Converting Date to DD-MM-YYYY and calculating Shift Time...")
        df_db['shift_date'] = df_db['shift_date'].apply(convert_date)
        df_db['pickup_time'] = df_db['pickup_time'].apply(convert_time)

        # 3. Logic: SHIFT TIME = PICKUP TIME + 2 HOURS
        # Convert DD-MM-YYYY back to datetime for calculation
        temp_pickup_dt = pd.to_datetime(df_db['shift_date'] + " " + df_db['pickup_time'], dayfirst=True, errors='coerce')
        
        # Add 2 Hours
        temp_shift_dt = temp_pickup_dt + pd.Timedelta(hours=2)
        shift_date_dt = pd.to_datetime(df_db["shift_date"], dayfirst=True).dt.date

        # 4. Populate Final Columns
        df_db['shift_time'] = temp_shift_dt.dt.strftime('%H:%M')
        fixed_drop_dt = temp_shift_dt.where(
            temp_shift_dt.dt.date == shift_date_dt,
            temp_shift_dt - pd.Timedelta(days=1)
        )
        df_db["drop_time"] = fixed_drop_dt.dt.strftime("%d-%m-%Y %H:%M")

        # Keep pickup_time as HH:MM
        pickup_dt = fixed_drop_dt - pd.Timedelta(hours=2)
        df_db["pickup_time"] = pickup_dt.dt.strftime("%d-%m-%Y %H:%M")


        # 5. Fill Excel Sheet cells with calculated data
        # We rewrite the Excel logic here to ensure the calculated fields are in the file
        for r_idx, row_data in enumerate(df_db.to_dict('records'), 2):
            for c_idx, header in enumerate(MANDATORY_HEADERS, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row_data.get(header, ""))
                cell.alignment = align_center
                cell.border = border

        # Headers to upper case
        # --- 3. THE CLEANING BLOCK (FIXED) ---
        # Convert all column names to UPPERCASE first
        df_db.columns = df_db.columns.str.strip().str.upper()

        if "EMPLOYEE_ADDRESS" in df_db.columns:
            df_db["EMPLOYEE_ADDRESS"] = (
                df_db["EMPLOYEE_ADDRESS"]
                .astype(str)
                .str.strip()
                .str.upper()
            )

        # Final conversion of all text to Upper
        text_cols = df_db.select_dtypes(include="object").columns
        df_db[text_cols] = df_db[text_cols].apply(lambda col: col.str.upper())
        df_db = df_db.fillna("").astype(str)

        # --- 4. WRITE CLEANED DATA TO EXCEL ---
        # Update headers to match uppercase
        FINAL_HEADERS = [h.upper() for h in MANDATORY_HEADERS]
        for col_idx, header in enumerate(FINAL_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)


        df_db = df_db.fillna("").astype(str)

    format_excel_sheet(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return df_db, output, "Operation_Cleaned.xlsx"
# ==========================================
# 2. MANUAL OPERATION DATA CLEANER
# ==========================================




def process_operaton_manual_data(file_data):
    """
    file_data = [(filename, bytes_content)]
    """

    final_dfs = []

    for file_name, content in file_data:

        # ------------------------------------------
        # 1. Read Excel (NO HEADER)
        # ------------------------------------------
        df = pd.read_excel(BytesIO(content), header=None)
        df = df.dropna(how="all")

        # ------------------------------------------
        # 2. Rename Columns by POSITION
        # ------------------------------------------
        new_columns = [
            "trip_id",          # 0
            "flight_number",    # 1
            "employee_id",      # 2
            "employee_name",    # 3
            "address",          # 4
            "contact_no",       # 5
            "cab_last_digit",   # 6
            "pickup_time",      # 7
            "reporting_time",   # 8
            "mis_remark"        # 9
        ]

        if len(df.columns) < len(new_columns):
            raise ValueError(
                f"{file_name}: Column count mismatch. Found {len(df.columns)}"
            )

        df = df.iloc[:, :len(new_columns)]
        df.columns = new_columns

        # ------------------------------------------
        # 3. Drop Invalid Employee Rows
        # ------------------------------------------
        

        # ------------------------------------------
        # 4. Extract Date from Filename
        # ------------------------------------------
        try:
            date_str = file_name.split()[-1].replace(".xlsx", "")
            file_date = datetime.strptime(date_str, "%d-%m-%Y").date()
        except Exception:
            file_date = None

        df["date"] = file_date

        # ------------------------------------------
        # 5. Reporting Area (Office) Logic
        # ------------------------------------------
        df["office"] = (
            df["address"]
            .astype(str)
            .str.extract(r'EMPLOYEE ADDRESS TO\s*"([^"]+)"', expand=False)
            .ffill()
        )
        df["office"] = df["office"].ffill()
        df = df[
            df["employee_id"].notna()
            & df["employee_id"].astype(str).str.strip().ne("")
            & df["employee_id"].astype(str).str.upper().ne("EMP ID")
        ]

        # ------------------------------------------
        # 6. Route No Formatting
        # ------------------------------------------
        df["trip_id"] = df["trip_id"].ffill()
        df["trip_id"] = "Route No.:- " + df["trip_id"].astype(str)

        final_dfs.append(df)

    # ------------------------------------------
    # 7. Merge All Files
    # ------------------------------------------
    if not final_dfs:
        return None, None, None

    final_df = pd.concat(final_dfs, ignore_index=True)

    # ------------------------------------------
    # 8. Export to BytesIO
    # ------------------------------------------
    output = BytesIO()
    final_df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    return final_df, output, "processed_operation_manual.xlsx"